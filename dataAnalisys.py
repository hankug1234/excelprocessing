import pandas as pd
import numpy as np
import csv
from openpyxl import load_workbook, Workbook
#"C:/Users/LG/Desktop/world-development-indicators/Indicators.xlsx""Indecators"

List = ["NY.ADJ.AEDU.GN.ZS",
"NY.GDP.MKTP.CD",
"SP.DYN.LE00.IN",
"SH.MED.CMHW.P3",
"SL.UEM.TOTL.ZS"]

keyvalue = [1,4]

def makeList(n,row):
    list = []
    for i in range(0,n):
        list.append(row[i].value)
    return list

def makeList2(n,row):
    list = []
    for i in range(0,n):
        list.append(row[i].value)
    return list

def attrSort(attrlist,keyvalue,filename,saveFileName,sheetName,indecate):
    wb = Workbook()
    ws = wb.active
    ws.title = sheetName

    indecator = load_workbook(filename=filename);
    sheet = indecator[sheetName]

    first_row = list(sheet.rows)[0]
    alist = []

    for i in range(0, len(keyvalue)):
        alist.append(first_row[keyvalue[i]].value)
    for i in range(0,len(attrlist)):
        alist.append(attrlist[i])
    ws.append(alist)

    wb.close()
    wb.save(saveFileName)

    indecator2 = load_workbook(filename=saveFileName);
    ws = indecator2[sheetName]
    sorce_rows = list(sheet.rows)[1:]
    dic = dict()
    count = 0
    t1 = sorce_rows[0][keyvalue[1]].value
    for k in sorce_rows:
        key1 = k[keyvalue[0]].value
        key2 = k[keyvalue[1]].value
        key0 = (str(key1)+str(key2))
        if( key0 in dic):
            dic[key0][k[indecate[0]].value] = k[indecate[1]].value
        else:
            dic[key0] = {}
            dic[key0][alist[0]] = key1
            dic[key0][alist[1]] = key2
            dic[key0][k[indecate[0]].value] = k[indecate[1]].value
        if(t1 != k[keyvalue[1]].value):
            dickey = dic.keys()
            appvalue = []
            for i in dickey:
                for j in alist:
                    if(j in dic[i] ):
                      appvalue.append(dic[i][j])
                    else:
                      appvalue.append("nan")
                ws.append(appvalue)
                del appvalue[:]
            dic = dict()
            indecator2.close()
            indecator2.save(saveFileName)
            indecator2 = load_workbook(filename=saveFileName);
            ws = indecator2[sheetName]
            t1 = k[keyvalue[1]].value

    dickey = dic.keys()
    appvalue = []
    for i in dickey:
        for j in alist:
            if (j in dic[i]):
                appvalue.append(dic[i][j])
            else:
                appvalue.append("nan")
        ws.append(appvalue)
        del appvalue[:]
    indecator2.close()
    indecator2.save(saveFileName)



def makeExcel(attr,filename,criteriarow,saveFileName,sheetName):
 wb = Workbook()
 ws = wb.active
 ws.title = sheetName

 indecator = load_workbook(filename=filename);
 sheet = indecator[sheetName]
 max_col = sheet.max_column


 wb.close()
 wb.save(saveFileName)
 count = 0;

 indecator2 = load_workbook(filename=saveFileName);
 ws = indecator2[sheetName]
 first = True
 for n in sheet.rows:
    if(first):
        ws.append(makeList(max_col, n))
        first = False

    if(n[criteriarow].value in attr):
        ws.append(makeList(max_col,n))
        count+=1

    if(count == 2000):
        indecator2.close()
        indecator2.save(saveFileName)
        count = 0;
        indecator2 = load_workbook(filename=saveFileName);
        ws = indecator2[sheetName]

 indecator2.close()
 indecator2.save(saveFileName)



'''
makeExcel(List,"C:/Users/LG/Desktop/world-development-indicators/Indicators.xlsx",3,"C:/Users/LG/Desktop/world-development-indicators/Indicators2.xlsx","Indicators")
attrSort(List,[1,4],"C:/Users/LG/Desktop/world-development-indicators/Indicators2.xlsx","C:/Users/LG/Desktop/world-development-indicators/Indicators3.xlsx","Indicators",[3,5])
'''
def attrCsv(attrlist,keyvalue,filename,saveFileName,indecate):
    raw = open(filename, 'r')
    wraw = open(saveFileName, 'a', newline='')
    writer = csv.writer(wraw)
    cooked = csv.reader(raw)
    for c in cooked:
     first_row = c
     break
    alist = []
    for i in range(0, len(keyvalue)):
        alist.append(first_row[keyvalue[i]])
    for i in range(0,len(attrlist)):
        alist.append(attrlist[i])
    writer.writerow(alist)

    sorce_rows = cooked
    dic = dict()
    for k in sorce_rows:
     t1 = k[keyvalue[1]]
     break

    for k in sorce_rows:
        key1 = k[keyvalue[0]]
        key2 = k[keyvalue[1]]
        key0 = (str(key1)+str(key2))
        if( key0 in dic):
            dic[key0][k[indecate[0]]] = k[indecate[1]]
        else:
            dic[key0] = {}
            dic[key0][alist[0]] = key1
            dic[key0][alist[1]] = key2
            dic[key0][k[indecate[0]]] = k[indecate[1]]
        if(t1 != k[keyvalue[1]]):
            dickey = dic.keys()
            appvalue = []
            for i in dickey:
                for j in alist:
                    if(j in dic[i] ):
                      appvalue.append(dic[i][j])
                    else:
                      appvalue.append("nan")
                writer.writerow(appvalue)
                del appvalue[:]
            dic = dict()
            t1 = k[keyvalue[1]]

    dickey = dic.keys()
    appvalue = []
    for i in dickey:
        for j in alist:
            if (j in dic[i]):
                appvalue.append(dic[i][j])
            else:
                appvalue.append("nan")
        writer.writerow(appvalue)
        del appvalue[:]

def makeCsv(filename,saveFileName,sheetName,criteriarow, attr):
 raw =  open(filename, 'r')
 wraw = open(saveFileName, 'a', newline='')
 writer = csv.writer(wraw)
 cooked = csv.reader(raw)

 wb = Workbook()
 ws = wb.active
 ws.title = sheetName

 count = 0;
 count2 = 0;
 first = True
 for n in cooked:
    count2+=1
    if (first):
        writer.writerow(n)
        first = False

    if (n[criteriarow] in attr):
        writer.writerow(n)
        count += 1

    if(count == 500000):
        wb.close()
        wb.save(saveFileName)
        count = 0;

    print(count2)




makeCsv("C:/Users/LG/Desktop/world-development-indicators/Indicators.csv","C:/Users/LG/Desktop/world-development-indicators/Indicators2.csv","Indicators",3,List)
attrCsv(List,[1,4],"C:/Users/LG/Desktop/world-development-indicators/Indicators2.csv","C:/Users/LG/Desktop/world-development-indicators/Indicators3.csv",[3,5])